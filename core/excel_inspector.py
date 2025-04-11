import os
import pandas as pd
from openpyxl import load_workbook

class ExcelFileInfo:
    """存储Excel文件的工作表和列信息"""
    def __init__(self, file_path):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.sheets = {}  # 工作表名称到列列表的映射

    def __repr__(self):
        return f"ExcelFileInfo('{self.file_name}', {len(self.sheets)} sheets)"

class ExcelInspector:
    """Excel文件检查器，用于获取Excel文件的工作表和列信息"""
    
    @staticmethod
    def get_excel_info(file_path):
        """获取Excel文件的工作表和列信息"""
        excel_info = ExcelFileInfo(file_path)
        
        try:
            # 使用openpyxl获取所有工作表名称
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            
            # 对每个工作表，获取列信息
            for sheet_name in sheet_names:
                # 使用pandas读取工作表的列名
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
                excel_info.sheets[sheet_name] = list(df.columns)
                
            return excel_info
            
        except Exception as e:
            # 如果使用openpyxl失败，尝试使用pandas
            try:
                # 获取所有工作表
                xlsx = pd.ExcelFile(file_path)
                sheet_names = xlsx.sheet_names
                
                # 对每个工作表，获取列信息
                for sheet_name in sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
                    excel_info.sheets[sheet_name] = list(df.columns)
                    
                return excel_info
                
            except Exception as e2:
                raise Exception(f"无法读取Excel文件信息: {str(e2)}")
    
    @staticmethod
    def batch_inspect_files(file_paths, progress_callback=None):
        """批量获取多个Excel文件的信息"""
        result = {}
        total_files = len(file_paths)
        
        for i, file_path in enumerate(file_paths):
            try:
                file_info = ExcelInspector.get_excel_info(file_path)
                result[file_path] = file_info
                
                # 如果有进度回调，则调用
                if progress_callback:
                    progress_percentage = int(((i + 1) / total_files) * 100)
                    progress_callback(progress_percentage, file_path)
                    
            except Exception as e:
                # 如果有进度回调，则调用
                if progress_callback:
                    progress_percentage = int(((i + 1) / total_files) * 100)
                    progress_callback(progress_percentage, file_path, str(e))
        
        return result 